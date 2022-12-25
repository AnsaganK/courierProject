from tempfile import NamedTemporaryFile

import openpyxl
from django.db.models import QuerySet
from django.http import HttpRequest, HttpResponse
from openpyxl.utils import get_column_letter

from app.models import Executor
from app.service.executor import get_query_parameters, _get_executor_json_list, get_filtered_executors
from app.service.period import get_last_periods


def generate_file_for_executors(request, executors: QuerySet):
    last_periods = get_last_periods()

    executors = get_filtered_executors(request=request, executors=executors).get('executors', [])
    columns = [
        '#', 'ID', 'Гражданство', 'ФИО', 'ЦФЗ', 'Номер', 'Рекрутер'
    ]
    columns += [str(period) for period in last_periods]
    columns += ['Активных часов']
    print(columns)
    executors = _get_executor_json_list(executors, last_periods)

    wb = create_file_for_export(columns, executors)
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    response = HttpResponse(content=stream, content_type='application/ms-excel', )
    response['Content-Disposition'] = f'attachment; filename=ExportedExcel.xlsx'
    return response


def create_file_for_export(columns, executors):
    wb = openpyxl.Workbook()
    executors_page = wb.active

    # header_columns_style =

    for index, column in enumerate(columns, start=1):
        cell = executors_page.cell(row=1, column=index)
        column_letter = get_column_letter(index)
        if column in ['ФИО', 'ЦФЗ']:
            width = 30
        elif column in ['#', 'Номер']:
            width = 10
        else:
            width = 20
        executors_page.column_dimensions[column_letter].width = width
        cell.value = column
    for index_executor, executor in enumerate(executors, start=2):
        hours = executor.get('hours')
        for index_colum, column in enumerate(columns, start=1):
            column_name = columns[index_colum - 1]
            cell = executors_page.cell(row=index_executor, column=index_colum)
            if column_name == '#':
                cell.value = index_executor - 1
            elif column_name == 'ID':
                cell.value = executor.get('executor_id')
            elif column_name == 'Гражданство':
                cell.value = executor.get('citizenship')
            elif column_name == 'ФИО':
                cell.value = executor.get('full_name', '')
            elif column_name == 'ЦФЗ':
                cell.value = executor.get('OFC')
            elif column_name == 'Номер':
                cell.value = executor.get('phone_number')
            elif column_name == 'Рекрутер':
                cell.value = executor.get('curator')
            elif column_name == 'Активных часов':
                cell.value = executor.get('hours_sum')
            else:
                for hour in hours:
                    if hour.get('period') == column_name:
                        cell.value = hour.get('hour')
    return wb


def create_sheet(sheet_name: str):
    pass


def create_cell(sheet):
    pass
