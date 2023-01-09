from datetime import datetime
from multiprocessing import Value

import pytz
from django.db.models import F
from django.db.models.functions import Concat
from openpyxl.reader.excel import load_workbook

from app.models import Executor, ArchiveFile, StatusChoices
from app.service.file import _get_columns
from utils import set_status


def set_executor_internship_dates(archive_file_id: int):
    archive_file = ArchiveFile.objects.get(pk=archive_file_id)
    set_status(archive_file, StatusChoices.WAIT)
    try:
        workbook = load_workbook(archive_file.file)
        worksheet = workbook.active
        columns = ['Город',
                   'ФИО',
                   'Телефон',
                   'Партнер',
                   'ЦФЗ стажировки',
                   'Основной ЦФЗ',
                   'Дата стажировки',
                   'Время стажировки',
                   'Статус стажировки',
                   'Причина отказа',
                   'Роль стажера',
                   'Дата создания стажировки',
                   'Дата изменения']
        columns = _get_columns(worksheet)

        max_row = worksheet.max_row
        for row in range(2, max_row + 1):
            executors = None
            date = None
            time = None
            for column_letter in columns:
                column_value = columns[column_letter].get('value')
                cell_value = worksheet[f'{column_letter}{row}'].value
                if column_value == 'ФИО':
                    full_name = cell_value.split(' ')
                    # print(full_name)
                    if len(full_name) > 2:
                        executors = Executor.objects.filter(
                            last_name=full_name[0],
                            first_name=full_name[1],
                            patronymic=' '.join(full_name[2:]),
                        )
                    elif len(full_name) == 2:
                        executors = Executor.objects.filter(
                            last_name=full_name[0],
                            first_name=full_name[1]
                        )
                    else:
                        executors = Executor.objects.annotate(
                            last_name=full_name[0]
                        )
                if column_value == 'Дата стажировки':
                    date = cell_value
                if column_value == 'Время стажировки':
                    time = cell_value
            if executors.exists() and date and time:
                date = datetime.strptime(date, '%d/%m/%y')
                time = datetime.strptime(time, '%H:%M')
                dt = datetime(year=date.year, month=date.month, day=date.day, hour=time.hour, minute=time.minute,
                              tzinfo=pytz.UTC)
                # print(dt)
                executors.update(internship_date=dt)
        set_status(archive_file, StatusChoices.SUCCESS)
    except Exception as e:
        print(e.__class__.__name__)
        print(e)
        set_status(archive_file, StatusChoices.ERROR)
