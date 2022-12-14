import datetime
from datetime import datetime

import openpyxl
from django.db.models import F, Value
from django.db.models.functions import Concat
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from app.models import Executor, CitizenshipType, Citizenship, OFC, ExecutorHours, ArchiveFile, Period, Day, DayHour, \
    StatusChoices, Transport, Contact
from app.service.city import add_city_for_ofc
from app.service.executor import add_ofc_for_executor
from utils import set_status


def get_file_data(file):
    workbook = load_workbook(file)
    worksheet = workbook.active
    columns = _get_columns(worksheet)
    print('Columns: ', columns)
    data = _get_data(worksheet, columns)
    return data


#
#                           Executor files tasks
#
def set_executors_file_data(archive_file_id):
    archive_file = ArchiveFile.objects.get(pk=archive_file_id)
    set_status(archive_file, StatusChoices.WAIT)
    try:
        file = archive_file.file
        workbook = load_workbook(file)
        worksheet = workbook.active
        columns = _get_columns(worksheet)
        _set_executor_data(worksheet, columns)
        set_status(archive_file, StatusChoices.SUCCESS)
    except Exception as e:
        print(e.__class__.__name__)
        print(e)
        set_status(archive_file, StatusChoices.ERROR)


def _get_columns(worksheet, row: int = 1) -> dict:
    max_column = worksheet.max_column

    columns = {}
    for column_index in range(1, max_column + 1):
        letter = get_column_letter(column_index)
        value = worksheet[f'{letter}{row}'].value
        if value:
            columns[letter] = {
                'index': column_index,
                'value': value
            }
    return columns


def _get_data(worksheet, columns):
    data = []
    max_row = worksheet.max_row
    for row in range(2, max_row + 1):
        data_item = {}
        for column_letter in columns:
            value = columns[column_letter].get('value')
            cell_value = worksheet[f'{column_letter}{row}'].value
            data_item[value] = cell_value
        data.append(data_item)
    print(data)
    return data


# Refactor
def _set_executor_data(worksheet, columns):
    max_row = worksheet.max_row
    for row in range(2, max_row + 1):
        executor = {}
        for column_letter in columns:
            value = columns[column_letter].get('value')
            cell_value = worksheet[f'{column_letter}{row}'].value
            _check_and_save_executor_attribute(executor, value, cell_value)
        _save_executor(executor)


def _save_executor(data: dict):
    executor = Executor.objects.get_or_create(executor_id=data.get('executor_id'))
    if executor[1]:
        executor[0].save()
    executor = executor[0]
    del data['executor_id']
    Executor.objects.filter(pk=executor.id).update(**data)


# Refactor
def _check_and_save_executor_attribute(executor: dict, value, cell_value):
    columns = ['????????????????????????', '??????', '??????????????', '??????', '????????????????', 'ID', '????????', '????????????????????', '???????????????? ??????????????',
               '???????? ?????????????????????? ????????????????', '???????? ???????????????????? ????????????????', '??????????????', '?????? ??????????????????????', '??????????????',
               '???? ??????????',
               '???????? ??????????????????????', '??????????????????????', '???????? ????????????????', '??????????????????????', '??????', '?????????? ?? ?????????? ????????????????',
               '???????? ???????????? ????????????????', '?????????? ???????????? ????????????????', '???????????????????????? ???????????? COVID19',
               '???????? ???????????????????? COVID19',
               '???????????? ???????? ???????????????????? COVID19', '????????????????', '???????? ?????????????????? ??????????????????', '???????????????????? ????????', '??????',
               '??????????????????',
               '??????????????????????????']

    if '????????????????????????' in value:
        pass
        # executor['full_name'] = cell_value
    if '??????????????' in value:
        executor['last_name'] = cell_value
    if '??????' in value:
        executor['first_name'] = cell_value
    if '????????????????' in value:
        executor['patronymic'] = cell_value
    if 'ID' == value:
        print(cell_value)
        executor['executor_id'] = cell_value
    if '????????' in value:
        transport = Transport.objects.get_or_create(name=cell_value)
        if transport[1]:
            transport[0].save()
        transport = transport[0]
        executor['transport'] = transport

    if '????????????????????' in value:
        executor['is_terminated'] = True if '????' in cell_value.lower() else False
    if '???????????????? ??????????????' in value:
        executor['main_contract'] = cell_value
    if '???????? ?????????????????????? ????????????????' in value and cell_value:
        if type(cell_value) == str:
            executor['date_terminated'] = datetime.strptime(cell_value, '%d.%m.%Y')
        elif type(cell_value) == datetime:
            executor['date_terminated'] = cell_value
    if '???????? ???????????????????? ????????????????' in value:
        if type(cell_value) == str:
            executor['date_conclusion'] = datetime.strptime(cell_value, '%d.%m.%Y')
        elif type(cell_value) == datetime:
            executor['date_conclusion'] = cell_value
    if '??????????????' in value:
        executor['partner'] = cell_value
    if '?????? ??????????????????????' in value:
        executor['citizenship_type'] = CitizenshipType.objects.filter(
            name__icontains='????????').first() if cell_value == 1 else CitizenshipType.objects.exclude(
            name__icontains='????????').first()
    if '??????????????' in value:
        executor['phone_number'] = cell_value
    if '???? ??????????' in value and cell_value:
        executor['email'] = cell_value.strip()
    if '???????? ??????????????????????' in value and cell_value:
        if type(cell_value) == str:
            executor['med_exam_date'] = datetime.strptime(cell_value, '%d.%m.%Y')
        elif type(cell_value) == datetime:
            executor['med_exam_date'] = cell_value
    if '??????????????????????' in value and cell_value:
        citizenship = Citizenship.objects.get_or_create(name=cell_value)
        if citizenship[1]:
            citizenship[0].save()
        citizenship = citizenship[0]
        executor['citizenship'] = citizenship

    if '???????? ????????????????' in value and cell_value:
        if type(cell_value) == str:
            executor['birth_date'] = datetime.strptime(cell_value, '%d.%m.%Y')
        elif type(cell_value) == datetime:
            executor['birth_date'] = cell_value

    if '??????????????????????' in value:
        executor['education'] = cell_value
    if '??????' in value and cell_value:
        executor[
            'gender'] = Executor.GenderChoices.MALE if '??????????????' in cell_value.lower() else Executor.GenderChoices.FEMALE
    if '?????????? ?? ?????????? ????????????????' in value and cell_value:
        executor['passport_series'] = str(cell_value).strip() if cell_value else str(cell_value)
    if '???????? ???????????? ????????????????' in value and cell_value:
        if type(cell_value) == str:
            executor['passport_date'] = datetime.strptime(cell_value, '%d.%m.%Y')
        elif type(cell_value) == datetime:
            executor['passport_date'] = cell_value
    if '?????????? ???????????? ????????????????' in value:
        executor['passport_place'] = cell_value
    if '???????????????????????? ???????????? COVID19' in value:
        pass
        # executor['covid19'] = cell_value
    if '???????????? ???????? ???????????????????? COVID19' in value:
        pass
        # executor['covid19_2'] = cell_value
    if '????????????????' in value:
        pass
        # executor['medotvod'] = cell_value

    if '???????? ?????????????????? ??????????????????' in value:
        pass
        # executor['date_medotvod'] = cell_value
    if '???????????????????? ????????' in value:
        executor['individual'] = cell_value
    if '??????' in value:
        executor['INN'] = cell_value
    if '??????????????????' in value:
        executor['note'] = cell_value
    if '??????????????????????????' in value and cell_value:
        ofc = OFC.objects.get_or_create(address=cell_value)
        if ofc[1]:
            ofc[0].save()
            add_city_for_ofc(ofc=ofc[0], many=False)
        ofc = ofc[0]
        executor['OFC'] = ofc


#
#                             Executor Hours files tasks
#
def set_executor_hours_file_data(archive_file_id: int):
    archive_file = ArchiveFile.objects.get(pk=archive_file_id)
    set_status(archive_file, StatusChoices.WAIT)
    try:
        workbook = load_workbook(archive_file.file)
        worksheet = workbook.active

        max_row = worksheet.max_row
        for row in range(1, max_row + 1):
            period_cell_value = worksheet[f'C{row}'].value
            table_first_cell_value = worksheet[f'A{row}'].value
            period_word = '????????????:'
            table_first_word = '??????????????????????????'
            if period_cell_value and period_word in period_cell_value:
                period_cell_value = period_cell_value.replace(period_word, '').replace(' ', '').strip().split('-')
                date_format = '%d.%m.%Y'
                print(period_cell_value)
                start_date = datetime.strptime(period_cell_value[0], date_format)
                final_date = datetime.strptime(period_cell_value[1], date_format)
                print(start_date)
                print(final_date)
                period = Period.objects.get_or_create(start_date=start_date, final_date=final_date)
                if period[1]:
                    period[0].save()
                period = period[0]
                print(period)

            if table_first_cell_value and table_first_word in table_first_cell_value:
                columns = _get_columns(worksheet, row=row)
                print(columns)
                _set_executor_hours_data(worksheet, row, columns, archive_file=archive_file, period=period)
        add_ofc_for_executor(many=True)

        set_status(archive_file, StatusChoices.SUCCESS)
    except:
        set_status(archive_file, StatusChoices.ERROR)


def _set_executor_hours_data(worksheet, row, columns, archive_file: ArchiveFile, period: Period):
    max_row = worksheet.max_row
    print('Row: ', row)
    for row in range(row + 2, max_row):
        executor_hour = {
            'archive_file': archive_file,
            'period': period,
            'hours': []
        }
        for column_letter in columns:
            value = columns[column_letter].get('value')
            cell_value = worksheet[f'{column_letter}{row}'].value
            _check_and_save_executor_hours_attribute(executor_hour, value, cell_value, period)
        print(executor_hour)
        _save_executor_hour(executor_hour)


def _check_and_save_executor_hours_attribute(executor_hour, value, cell_value, period):
    if '??????????????????????????' in value:
        ofc = OFC.objects.get_or_create(address=cell_value)
        if ofc[1]:
            ofc[0].save()
            add_city_for_ofc(ofc=ofc[0], many=False)
        ofc = ofc[0]
        executor_hour['ofc'] = ofc
    elif '??????????????????????' in value:
        executor_hour['executor_fullname'] = cell_value
    elif '????????' in value:
        executor_hour['role'] = cell_value
        transport = Transport.objects.get_or_create(name=cell_value)
        if transport[1]:
            transport[0].save()
        transport = transport[0]
        executor_hour['transport'] = transport
    elif cell_value and 'ID' in value:
        executor = Executor.objects.get_or_create(executor_id=cell_value)
        if executor[1]:
            executor[0].save()
        executor = executor[0]
        print(executor.get_full_name, executor_hour.get('executor_fullname'))
        if executor.get_full_name == '-' and executor_hour.get('executor_fullname'):
            fullname = executor_hour.get('executor_fullname').split(' ')
            print(fullname)
            if len(fullname) >= 1:
                executor.last_name = fullname[0]
            if len(fullname) >= 2:
                executor.first_name = fullname[1]
            if len(fullname) >= 3:
                executor.patronymic = ' '.join(fullname[2:])
        if not executor.transport and executor_hour.get('transport'):
            executor.transport = executor_hour.get('transport')
            executor.save()

        executor_hour['executor'] = executor
        executor_hour['executor_id'] = cell_value
    else:
        try:
            day = datetime.strptime(value, '%d.%m.%y')
            day = Day.objects.get_or_create(date=day, period=period)
            if day[1]:
                day[0].save()
            day = day[0]

            executor_hour['hours'].append({
                'hour': float(cell_value) if cell_value else 0.0,
                'day': day
            })
        except Exception as e:
            pass


def _save_executor_hour(data: dict):
    for hour_object in data['hours']:
        transport = Transport.objects.get_or_create(name=data.get('transport'))
        if transport[1]:
            transport[0].save()
        transport = transport[0]
        executor = data.get('executor')
        executor_hour = ExecutorHours.objects.get_or_create(
            executor=executor,
            ofc=data.get('ofc'),
            transport=transport,
            period=data.get('period'),
        )

        if executor_hour[1]:
            executor_hour[0].save()
        executor_hour = executor_hour[0]
        executor_hour.files.add(data.get('archive_file'))
        executor_hour.save()

        day_hour = DayHour.objects.get_or_create(executor_hour=executor_hour, day=hour_object['day'])
        if day_hour[1]:
            day_hour[0].save()
        day_hour = day_hour[0]
        day_hour.hour = hour_object.get('hour')
        day_hour.save()

    executor = data.get('executor')
    if executor:
        executor.OFC = data.get('ofc')
        executor.save()
    # ExecutorHours.objects.filter(pk=executor_hour.id).update(hour=hour_object['hour'])


#
#                                 Executor Phones File Task
#
def set_executor_phones_file_data(archive_file_id: int):
    archive_file = ArchiveFile.objects.get(pk=archive_file_id)
    set_status(archive_file, StatusChoices.WAIT)
    try:
        workbook = load_workbook(archive_file.file)
        worksheet = workbook.active

        max_row = worksheet.max_row
        for row in range(1, max_row + 1):
            fullname = worksheet[f'A{row}'].value
            fullname = fullname.split(' ')
            fullname = [i for i in fullname if i != '']
            last_name = None
            first_name = None
            patronymic = None
            if len(fullname) >= 1:
                last_name = fullname[0]
            if len(fullname) >= 2:
                first_name = fullname[1]
            if len(fullname) >= 3:
                patronymic = ' '.join(fullname[2:])
            number = worksheet[f'B{row}'].value
            executors = Executor.objects.filter(last_name=last_name, first_name=first_name, patronymic=patronymic)
            if executors.exists():
                for executor in executors:

                    contacts = Contact.objects.filter(
                        executor=executor,
                        type=Contact.TypeChoices.WHATSAPP
                    )
                    if contacts.exists():
                        contact = contacts.first()
                        contact.identifier = number
                    else:
                        contact = Contact.objects.create(
                            identifier=number,
                            type=Contact.TypeChoices.WHATSAPP,
                            executor=executor)
                    contact.save()

                    executor.phone_number = number
                    executor.save()
                print(f'{executors.count()} {fullname} {number}')
            else:
                # print(f'{last_name}, {first_name}, {patronymic}')
                print(f'NONE {fullname} {number}')
        set_status(archive_file, StatusChoices.SUCCESS)
    except Exception as e:
        print(e)
        print(e.__class__.__name__)
        set_status(archive_file, StatusChoices.ERROR)
